#!  /usr/bin/env python
#ecoding=utf-8

import csv
import pandas
import openpyxl as pxl


#方式一csv
def group_history2excel():
    result=[]
    with open('room_history.txt', 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip().split('##$##')
            roomname = line[0]
            username =  line[1]
            text = line[2]
            msgtime =  line[3]
            result.append([msgtime,text,username,roomname])
    with open('Result.csv', 'w',encoding='utf_8_sig',newline='') as f:
        writer = csv.writer(f,dialect='excel')
        writer.writerow(['时间','内容', '用户', '群名'])
        writer.writerows(result)

#方式二pandas
def group_history2excel_pandas():
    result=[['时间', '内容', '用户', '群名']]
    with open('room_history.txt', 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip().split('##$##')
            if len(line)>=4:
                roomname = line[0]
                username =  line[1]
                text = line[2]
                msgtime =  line[3]
                result.append([msgtime,text,username,roomname])
            else:
                print('该数据不完整!\n'+str(line))
    #save to excel
    data=pandas.DataFrame(result)
    writer=pandas.ExcelWriter('Result.xlsx')
    data.to_excel(writer)
    writer.save()

#方式三openpyxl
def save2excel():
    wb = pxl.Workbook()#就新建了一个新的工作表
    sheet=wb.create_sheet('sheet1', index=0) # 被安排到第二个工作表，index=0就是第一个位置
    result=['时间', '内容', '用户', '群名']
    sheet.append(result)
    wb.save(r'Result.xlsx')


#openpyxl 读取excel
def read_excel2():
    wb=pxl.load_workbook("Result.xlsx")
    sheet_name=wb.get_sheet_names()[0]
    sheet=wb.get_sheet_by_name(sheet_name)
    # 因为按行，所以返回A1, B1, C1这样的顺序
    for row in sheet.rows:
        for cell in row:
            print(cell.value)
    # A1, A2, A3这样的顺序
    for column in sheet.columns:
        for cell in column:
            print(cell.value)

#pandas 读取excel
def read_excel(path):
    xls_file=pandas.ExcelFile(path)
    for name in xls_file.sheet_names:#显示出读入excel文件中的表名字
        print('+++++++++++++++++++++++++++++sheet name:'+name+'+++++++++++++++++++++++++++++')
        sheet1=xls_file.parse(name)
        col_name=sheet1.keys()[1]#每一列的第一行为键
        col_vals=[col_name]
        for col_val in sheet1.pop(col_name):
            col_vals.append(col_val)
        print(col_vals)



if __name__=='__main__':
    # group_history2excel();
    # group_history2excel_pandas()
    # read_excel('采集频道总表.xlsx')
    # save2excel()
    read_excel2()