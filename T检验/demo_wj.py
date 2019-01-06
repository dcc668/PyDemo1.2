import openpyxl as pxl
import re
from scipy import stats
import csv
import  numpy as np



def test_reader():
    #依次从大一（身高体重肺活量）  男生和女生
    # 其中a代表体侧项目，从0开始
    idx = 0
    list_count = []
    projects={}#存所有项目的 所有年级的 男和女
    for a in range(10):
        print("体侧项目：",a)
        one2four=[]#a项目的 所有年级的 男和女
        for b in range(4):
            print("年级名称：", b)
            list1, list2 = get_data_from_xlsx(a+1, b+1)
            one2four.append(list1)#b年级  男
            one2four.append(list2)#b年级  女
        projects[a]=one2four
    print(projects);
    return projects;

def get_data_from_xlsx(item=1, grade=1):
    wb = pxl.load_workbook("2017.xlsx")

    sheet1 = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(sheet1)

    list_male = []
    list_female = []

    # 计算男女生合计项目成绩
    if grade == 5:
        for i in range(3):
            list_male.append(ws.cell(row=1 + 4 * grade, column=item * 3 + i).value)
            #list_female.append(ws.cell(row=1 + 4 * grade, column=item * 3 + i).value)

        return(list_male, list_male)


    # 800米项目和1000米项目比较
    if item == 6:
        for i in range(3):
            '''# 根据单元格划分，规定身高、大一为（1，1）
            按照行取数据，row=1代表男生;按照项目取item，column=item*3一个项目有3个单元格子；
            依次循环遍历，年级、项目'''
            '''根据excel表格条件需要，男生800米项目和女生1000米项目成绩相比较'''

            list_male.append(ws.cell(row=1 + 4 * grade, column=(item + 1) * 3 + i).value)
            list_female.append(ws.cell(row=3 + 4 * grade, column=item * 3 + i).value)
    # 1000米项目和800米项目男女生比较
    elif item == 7:
        for i in range(3):
            list_male.append(ws.cell(row=1 + 4 * grade, column=item * 3 + i).value)
            list_female.append(ws.cell(row=3 + 4 * grade, column=(item - 1) * 3 + i).value)
    # 引体向上和一分钟仰卧起做比较
    elif item == 8:
        for i in range(3):
            list_male.append(ws.cell(row=1 + 4 * grade, column=item * 3 + i).value)
            list_female.append(ws.cell(row=3 + 4 * grade, column=(item + 2) * 3 + i).value)
    # 一分钟仰卧起做和引起向上比较
    elif item == 10:
        for i in range(3):
            list_male.append(ws.cell(row=1 + 4 * grade, column=(item - 2) * 3 + i).value)
            list_female.append(ws.cell(row=3 + 4 * grade, column=item * 3 + i).value)
    # 依次遍历每个项目
    else:
        for i in range(3):
            '''# 根据单元格划分，规定身高、大一为（1，1）
            按照行取数据，row=1代表男生;按照项目取item，column=item*3一个项目有3个单元格子；
            依次循环遍历，年级、项目'''
            list_male.append(ws.cell(row=1 + 4 * grade, column=item * 3 + i).value)
            list_female.append(ws.cell(row=3 + 4 * grade, column=item * 3 + i).value)

    return (list_male, list_female)

def replaceNullStr(old):
    if old!='':
        return float(str(old).replace("‘","."))
    else:
        return 0.0001


if __name__ == "__main__":

    projects=test_reader();
    sp=[]
    for i in range(10):# 按体育项目 遍历
        items=projects[i];
        if i==6:#800米跑
            projects[7]#1000米跑
        elif i!=7:#除了 800,1000米跑的体育项目
            cols=[]#一个体育项目的比较结果  放到一列
            for j in range(len(items)-1):#冒泡 取a，b 调用T检验
                for k in range(j,len(items)):
                    a=items[j]
                    b=items[k]
                    me1=replaceNullStr(a[1])
                    me2=replaceNullStr(b[1])
                    st1=replaceNullStr(a[2])
                    st2=replaceNullStr(b[2])
                    no1=replaceNullStr(a[0])
                    no2=replaceNullStr(b[0])
                    # print(str(me1)+"-"+str(me2)+"-"+str(st1)+"-"+str(st2)+"-"+str(no1)+"-"+str(no2))
                    if no1==0 or no2==0:#样本数为0 包除0异常
                        cols.append("0.0/0.0")
                        continue
                    res=stats.ttest_ind_from_stats(mean1=me1, std1=st1, nobs1=no1,
                                                   mean2=me2, std2=st2, nobs2=no2)

                    cols.append(str(res.statistic)+"/"+str(res.pvalue))
            print("列数据 个数："+str(len(cols)))
            sp.append(cols)
    print("行数据 个数："+str(len(sp)))
    print(sp)
    np.reshape(sp,(-1,1))
    new_cols=[]
    for line in sp:
        # col=np.reshape(line,(-1,1)).T#列变行 变成一列
        for it in line:
            new_cols.append(it)
    result=np.reshape(new_cols,(-1,8)).T#列变行 8列变8行
    print(result)




        # # list1, list2 = get_data_from_xlsx(6, 2)
    #
    # result = []
    # dayinan, dayinv = get_data_from_xlsx(1, 1)
    # print(dayinan)
    # print(dayinv)
    # #result.append(stats.ttest_ind_from_stats(mean1=dayinan[1], std1=dayinan[2], nobs1=dayinan[0],
    # #                                         mean2=dayinan[1], std2=dayinan[2], nobs2=dayinan[0]))
    #
    # # result.append(stats.ttest_ind_from_stats(mean1=dayinan[1], std1=dayinan[2], nobs1=dayinan[0],
    # #                                              mean2=dayinv[1], std2=dayinv[2], nobs2=dayinv[0]))
    # res=stats.ttest_ind_from_stats(mean1=dayinan[1], std1=dayinan[2], nobs1=dayinan[0],
    #                        mean2=dayinv[1], std2=dayinv[2], nobs2=dayinv[0])
    # result.append([res.statistic,res.pvalue])
    # print(result)
    # '''# 想法是依次遍历每一项目，例如身高：大一男跟身高项目中所有数据对比，包括他自己；
    # 根据T检验结果输出T矩阵'''
    # for j in range(4):
    #     dayinan, dayinv = get_data_from_xlsx(1, j)
    #     for i in range(4):
    #         list_nan, list_nv = get_data_from_xlsx(1, i)
    #
    #         print(list_nan)
    #         print(list_nv)
    #         # T检验函数，其中mean代表均值，srd代表方差，nobs代表样本数
    #         result.append(stats.ttest_ind_from_stats(mean1=dayinan[1], std1=dayinan[2], nobs1=dayinan[0],
    #                                                  mean2=list_nan[1], std2=list_nan[2], nobs2=list_nan[0]))
    #
    #         result.append(stats.ttest_ind_from_stats(mean1=dayinan[1], std1=dayinan[2], nobs1=dayinan[0],
    #                                                  mean2=list_nv[1], std2=list_nv[2], nobs2=list_nv[0]))


    # for
    #
    #
    #
    # for i in range(len(result)):
    #     print(result[i])
    #
    #
    # with open("demo.csv", 'w', newline='') as csvfile:
    #     csv_writer = csv.writer(csvfile)
    #     csv_writer.writerow()


    #test_reader()
    # get_data_from_xlsx()
    # test_reader()
    # print(list1)
    # print(list2)
    # print("++++++++++++++++++")