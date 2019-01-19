#ecoding=utf-8
import pandas
import  requests
from lxml import html
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from requests import Session
from selenium.webdriver import Chrome,PhantomJS
import time,json,random
from settings import AGENTS
from openpyxl.reader.excel import load_workbook
import openpyxl as pxl
class ZhiHuSelenum():
    def __init__(self):
        self.tomJsDriver=r'E:\pycharm_space\Demo1.2\driver\phantomjs.exe';
        self.chromeDriver = r'E:\pycharm_space\Demo1.2\driver\chromedriver.exe'
        self.session = Session()
        self.session.headers.clear()

    #工具方法
    #tomjs初始化
    def tomjs_init(self):
        # 引入配置对象DesiredCapabilities
        dcap = dict(DesiredCapabilities.PHANTOMJS)
        # 从USER_AGENTS列表中随机选一个浏览器头，伪装浏览器
        agent=random.choice(AGENTS);
        dcap["phantomjs.page.settings.userAgent"] =agent
        dcap["phantomjs.page.customHeaders.User-Agent"] = agent
        # 不载入图片，爬页面速度会快很多
        dcap["phantomjs.page.settings.loadImages"] = False
        # #打开带配置信息的phantomJS浏览器
        tomJs = PhantomJS(self.tomJsDriver, desired_capabilities=dcap,service_args=['--ignore-ssl-errors=true', '--ssl-protocol=TLSv1'])
        # 隐式等待5秒，可以自己调节
        tomJs.implicitly_wait(10)
        # 设置10秒页面超时返回，类似于requests.get()的timeout选项，driver.get()没有timeout选项
        # 以前遇到过driver.get(url)一直不返回，但也不报错的问题，这时程序会卡住，设置超时选项能解决这个问题。
        tomJs.set_page_load_timeout(10)
        # 设置10秒脚本超时时间
        tomJs.set_script_timeout(10)
        return tomJs
    def startRun(self):
        tomJs=self.tomjs_init();
        # title = '院校名称 院校所在地 院校隶属 院校类型 学历层次 985/211  满意度'.split(" ")
        # wb = pxl.Workbook()  # 就新建了一个新的工作表
        # sheet = wb.create_sheet('sheet1', index=0)  # 被安排到第二个工作表，index=0就是第一个位置
        # sheet.append(title)
        # wb.save(r'Result.xlsx')
        # wb.close()
        for i in range(127,138):
            url="https://gaokao.chsi.com.cn/sch/search--ss-on,searchType-1,option-qg,start-"+str(i*20)+".dhtml"
            print("request:"+url)
            tomJs.get(url)
            try:
                trs=tomJs.find_elements_by_xpath("//table[@class='ch-table']/tbody/tr")
                result2 = []
                for tr in trs[1:]:
                    texts=tr.text
                    txx = texts.split(" ")
                    #去除空格
                    new_tx=[]
                    for t in txx:
                        if len(t)==1 or len(t)==0:
                            pass
                        else:
                            new_tx.append(t)
                    if(len(new_tx)==7):
                        new_tx.append(" ")
                        new_tx[7]=new_tx[6]
                        new_tx[6]=""
                    if (len(new_tx) == 6):
                        new_tx.append(" ")
                        new_tx.append(" ")
                        new_tx[7] = new_tx[5]
                        new_tx[5] = ""
                    result2.append(new_tx)
                self.save2excel(result2)
                print(str(i)+"获取学校个数:" + str(len(result2)))
                time.sleep(2)
            except Exception as e:
                print(e)
                return;


    def save2excel(self, result):
        wb = load_workbook(filename=r'Result.xlsx')
        sheetnames = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheetnames[0])  # 取第一张表
        for res in result:
            sheet.append(res)
        wb.save(r'Result.xlsx')
        wb.close()
if __name__=='__main__':
    zhihu=ZhiHuSelenum()
    zhihu.startRun()
